from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from sbi_fund_faq.ingestion.models import IngestedChunk, clean_text
from sbi_fund_faq.ingestion.sources import SourceDocument


def parse_ter_file(source: SourceDocument, source_path: Path) -> list[IngestedChunk]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    worksheet = workbook["Sheet1"]
    rows = worksheet.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(rows)]
    supported_schemes = {scheme.lower(): scheme for scheme in source.scheme_names}
    chunks: list[IngestedChunk] = []

    for row_number, row in enumerate(rows, start=2):
        row_data = dict(zip(headers, row))
        raw_scheme_name = clean_text(row_data.get("Scheme Name"))
        scheme_name = supported_schemes.get(raw_scheme_name.lower())
        if not scheme_name:
            continue

        ter_date = normalize_date(row_data.get("TER Date (DD/MM/ YYYY)"))
        regular_total_ter = format_percent(row_data.get("Regular Plan - Total TER (%)"))
        direct_total_ter = format_percent(row_data.get("Direct Plan - Total TER (%)"))

        chunk_text = (
            f"{scheme_name} TER on {ter_date}: "
            f"Regular Plan Total TER {regular_total_ter}; "
            f"Direct Plan Total TER {direct_total_ter}."
        )

        chunks.append(
            IngestedChunk(
                id=f"{source.id}-{slug(scheme_name)}-row-{row_number}",
                source_name=source.source_name,
                scheme_name=scheme_name,
                document_type=source.document_type,
                document_month=source.document_month,
                page_number=1,
                section="Total Expense Ratio",
                chunk_text=chunk_text,
                source_id=source.id,
                file_name=source.file_name,
                metadata={
                    "sheet_name": worksheet.title,
                    "row_number": row_number,
                    "ter_date": ter_date,
                    "nsdl_scheme_code": clean_text(row_data.get("NSDL Scheme Code")),
                    "regular_plan_total_ter_percent": regular_total_ter,
                    "direct_plan_total_ter_percent": direct_total_ter,
                    "regular_plan_base_expense_ratio_percent": format_percent(
                        row_data.get("Regular Plan - Base Expense Ratio (BER) (%)")
                    ),
                    "direct_plan_base_expense_ratio_percent": format_percent(
                        row_data.get("Direct Plan - Base Expense Ratio (BER) (%)")
                    ),
                },
            )
        )

    workbook.close()
    return chunks


def normalize_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return clean_text(value)


def format_percent(value: Any) -> str:
    if isinstance(value, (int, float)):
        return f"{value:.2f}%"
    text = clean_text(value)
    if not text:
        return ""
    return text if text.endswith("%") else f"{text}%"


def slug(value: str) -> str:
    return "-".join(clean_text(value).lower().split())
